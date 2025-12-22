import xml.etree.ElementTree as ET
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum
import csv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class AcquisitionMode(Enum):
    """Camera acquisition modes."""
    VIS_SCIENCE = "AcquireVisCamScienceData"
    VIS_FFI = "AcquireVisCamImages"
    NIR = "AcquireInfCamImages"


@dataclass
class ObservationModes:
    """Tracks which modes are used in an observation sequence."""
    vis_mode: Optional[AcquisitionMode] = None  # Either SCIENCE or FFI
    has_nir: bool = False
    
    def __str__(self) -> str:
        modes = []
        if self.vis_mode:
            modes.append(self.vis_mode.value)
        if self.has_nir:
            modes.append(AcquisitionMode.NIR.value)
        return " + ".join(modes) if modes else "No acquisition modes"


@dataclass
class ObservationSequenceData:
    """Data for a single observation sequence."""
    obs_id: str
    sequence_id: str
    task_id: str
    target: Optional[str] = None
    modes: ObservationModes = field(default_factory=ObservationModes)
    vis_params: Dict[str, Any] = field(default_factory=dict)
    nir_params: Dict[str, Any] = field(default_factory=dict)
    data_volumes: Dict[str, float] = field(default_factory=dict)  # mode -> bytes


@dataclass
class VisitData:
    """Data for entire visit with all observation sequences."""
    visit_id: str
    observation_sequences: List[ObservationSequenceData] = field(default_factory=list)
    total_data_volumes: Dict[str, float] = field(default_factory=dict)  # mode -> total bytes
    
    def add_sequence(self, seq: ObservationSequenceData):
        """Add a sequence and accumulate data volumes."""
        self.observation_sequences.append(seq)
        for mode, volume in seq.data_volumes.items():
            if mode not in self.total_data_volumes:
                self.total_data_volumes[mode] = 0.0
            self.total_data_volumes[mode] += volume


@dataclass
class DataVolumeAlert:
    """Alert for sequences exceeding data volume threshold."""
    visit_id: str
    obs_id: str
    task_id: str
    mode: str
    volume_gb: float
    threshold_gb: float = 1.8
    
    def __str__(self) -> str:
        return f"⚠️  Task {self.task_id} - Visit {self.visit_id} - Sequence {self.obs_id} - {self.mode}: {self.volume_gb:.4f} GB (exceeds {self.threshold_gb} GB)"


class SOCXMLParser:
    """Parser for Pandora SOC calendar XML files."""
    
    # Data volume calculation constants
    BITS_PER_PIX_VIS = 32
    BITS_PER_PIX_NIR = 32
    COMPRESSION_RATIO_VIS = 0.3
    COMPRESSION_RATIO_NIR = 0.7
    OVERHEAD_BUFFER_RATIO = 1.1
    GENERAL_BUFFER_RATIO = 1.25
    FRAME_TIME_VIS = 0.2  # seconds
    DATA_VOLUME_THRESHOLD_GB = 1.8  # Flag threshold

    
    def __init__(self, xml_file_path: str):
        """
        Initialize parser with XML file path.
        
        Args:
            xml_file_path: Path to the SOC XML calendar file
        """
        self.xml_file_path = xml_file_path
        self.tree = ET.parse(xml_file_path)
        self.root = self.tree.getroot()
        self._strip_namespaces(self.root)
    
    def _strip_namespaces(self, elem: ET.Element):
        """Remove namespace prefixes from all tags."""
        if "}" in elem.tag:
            elem.tag = elem.tag.split("}", 1)[1]
        for child in elem:
            self._strip_namespaces(child)
    
    def parse_visit(self) -> VisitData:
        """
        Parse entire visit from XML.
        
        Returns:
            VisitData containing all observation sequences and data volumes
        """
        visit_elem = self.root.find(".//Visit")
        if visit_elem is None:
            raise RuntimeError("No <Visit> element found in XML")
        
        visit_id = self._get_text(visit_elem, "ID")
        visit = VisitData(visit_id=visit_id)
        
        # Parse all observation sequences in the visit
        seq_elems = visit_elem.findall("Observation_Sequence")
        for seq_elem in seq_elems:
            seq_data = self._parse_observation_sequence(seq_elem)
            visit.add_sequence(seq_data)
        
        return visit
    
    def parse_all_visits(self) -> Tuple[List[VisitData], List[DataVolumeAlert]]:
        """Parse every Visit in the XML and flag high-volume sequences."""
        visits: List[VisitData] = []
        alerts: List[DataVolumeAlert] = []
        visit_elems = self.root.findall(".//Visit")
        if not visit_elems:
            raise RuntimeError("No <Visit> elements found in XML")

        for visit_elem in visit_elems:
            visit_id = self._get_text(visit_elem, "ID")
            visit = VisitData(visit_id=visit_id)

            seq_elems = visit_elem.findall("Observation_Sequence")
            for seq_elem in seq_elems:
                seq_data = self._parse_observation_sequence(seq_elem)
                visit.add_sequence(seq_data)
                
                # Check for high-volume sequences
                for mode, vol_bytes in seq_data.data_volumes.items():
                    vol_gb = vol_bytes / (1e9)
                    if vol_gb > self.DATA_VOLUME_THRESHOLD_GB:
                        alerts.append(DataVolumeAlert(
                            visit_id=visit_id,
                            obs_id=seq_data.obs_id,
                            task_id=seq_data.task_id,
                            mode=mode,
                            volume_gb=vol_gb
                        ))

            visits.append(visit)
        return visits, alerts
    
    def _parse_observation_sequence(self, seq_elem: ET.Element) -> ObservationSequenceData:
        """
        Parse a single observation sequence.
        
        Args:
            seq_elem: XML element for Observation_Sequence
            
        Returns:
            ObservationSequenceData with all parameters and calculated volumes
        """
        obs_id = self._get_text(seq_elem, "ID")
        task_id = self._get_text(seq_elem, "Task")[:4]
        
        # Extract observational parameters
        obs_params = seq_elem.find("Observational_Parameters")
        target = self._get_text(obs_params, "Target") if obs_params else None
        
        seq_data = ObservationSequenceData(
            obs_id=obs_id,
            sequence_id=obs_id,
            task_id=task_id,
            target=target
        )
        
        # Parse payload parameters
        payload = seq_elem.find("Payload_Parameters")
        if payload is None:
            return seq_data
        
        # Check for VIS camera modes
        vis_science = payload.find("AcquireVisCamScienceData")
        vis_ffi = payload.find("AcquireVisCamImages")
        
        if vis_science is not None:
            seq_data.modes.vis_mode = AcquisitionMode.VIS_SCIENCE
            seq_data.vis_params = self._extract_all_params(vis_science)
            vis_vol = self._calculate_vis_science_volume(seq_data.vis_params)
            seq_data.data_volumes[AcquisitionMode.VIS_SCIENCE.value] = vis_vol
        elif vis_ffi is not None:
            seq_data.modes.vis_mode = AcquisitionMode.VIS_FFI
            seq_data.vis_params = self._extract_all_params(vis_ffi)
            vis_vol = self._calculate_vis_ffi_volume(seq_data.vis_params)
            seq_data.data_volumes[AcquisitionMode.VIS_FFI.value] = vis_vol
        
        # Check for NIR camera mode
        nir = payload.find("AcquireInfCamImages")
        if nir is not None:
            seq_data.modes.has_nir = True
            seq_data.nir_params = self._extract_all_params(nir)
            nir_vol = self._calculate_nir_volume(seq_data.nir_params)
            seq_data.data_volumes[AcquisitionMode.NIR.value] = nir_vol
        
        return seq_data
    
    def _extract_all_params(self, elem: ET.Element) -> Dict[str, Any]:
        """Extract all child elements as parameter dictionary."""
        params = {}
        for child in elem:
            text = child.text
            # Try to convert to numeric types
            if text:
                try:
                    if "." in text:
                        params[child.tag] = float(text)
                    else:
                        params[child.tag] = int(text)
                except ValueError:
                    params[child.tag] = text
            else:
                params[child.tag] = text
        return params
    
    def _get_text(self, elem: ET.Element, tag: str) -> Optional[str]:
        """Get text content of child element."""
        child = elem.find(tag)
        return child.text if child is not None and child.text else None
    
    def _calculate_vis_ffi_volume(self, params: Dict[str, Any]) -> float:
        """
        Calculate data volume for VIS FFI mode (AcquireVisCamImages).
        
        Formula: ROI_SizeX * ROI_SizeY * NumExposures * (32 bits/pix) * 
                 compression_ratio * overhead * general_buffer
        """
        roi_x = params.get("ROI_SizeX", 0)
        roi_y = params.get("ROI_SizeY", 0)
        num_exposures = params.get("NumExposures", 0)
        
        bytes_per_pix = self.BITS_PER_PIX_VIS / 8
        data_vol = (roi_x * roi_y * num_exposures * bytes_per_pix) * \
                   (self.COMPRESSION_RATIO_VIS * self.OVERHEAD_BUFFER_RATIO * 
                    self.GENERAL_BUFFER_RATIO)
        
        return data_vol
    
    def _calculate_vis_science_volume(self, params: Dict[str, Any]) -> float:
        """
        Calculate data volume for VIS Science mode (AcquireVisCamScienceData).
        
        Formula: StarRoiDimension^2 * MaxNumStarRois * NumTotalFramesRequested * 
                 (32 bits/pix) * compression_ratio * overhead * general_buffer
        """
        roi_dim = params.get("StarRoiDimension", 0)
        max_star_rois = params.get("MaxNumStarRois", 0)
        num_frames = params.get("NumTotalFramesRequested", 0) / params.get("FramesPerCoadd", 0)
        
        bytes_per_pix = self.BITS_PER_PIX_VIS / 8
        data_vol = (roi_dim * roi_dim * max_star_rois * num_frames * bytes_per_pix) * \
                   (self.COMPRESSION_RATIO_VIS * self.OVERHEAD_BUFFER_RATIO * 
                    self.GENERAL_BUFFER_RATIO)
        
        return data_vol
    
    def _calculate_nir_volume(self, params: Dict[str, Any]) -> float:
        """
        Calculate data volume for NIR mode (AcquireInfCamImages).
        
        Formula depends on AverageGroups:
        - If AverageGroups=0: SC_ReadFrames * SC_Groups * SC_Integrations
        - If AverageGroups=1: SC_Groups * SC_Integrations
        """
        roi_x = params.get("ROI_SizeX", 0)
        roi_y = params.get("ROI_SizeY", 0)
        avg_groups = params.get("AverageGroups", 0)
        sc_readframes = params.get("SC_ReadFrames", 1)
        sc_groups = params.get("SC_Groups", 1)
        sc_integrations = params.get("SC_Integrations", 0)
        
        bytes_in_frame = roi_x * roi_y * (self.BITS_PER_PIX_NIR / 8)
        
        if avg_groups == 0:
            total_frames = sc_readframes * sc_groups * sc_integrations
        else:  # AverageGroups == 1
            total_frames = sc_groups * sc_integrations
        
        data_vol_uncompressed = bytes_in_frame * total_frames
        data_vol = data_vol_uncompressed * (self.COMPRESSION_RATIO_NIR * 
                                             self.OVERHEAD_BUFFER_RATIO * 
                                             self.GENERAL_BUFFER_RATIO)
        
        return data_vol

def export_to_excel(visits: List[VisitData], alerts: List[DataVolumeAlert], output_file: str = "data_volumes.xlsx"):
    """
    Export data volumes to Excel with three sheets:
    1. Sequences: sequence-level data volumes
    2. Visits: visit-level totals
    3. Tasks: task-level totals (aggregated by task_id)
    
    Args:
        visits: List of VisitData objects
        alerts: List of DataVolumeAlert objects
        output_file: Output Excel file path
    """
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ===== SHEET 1: SEQUENCES =====
    ws_seq = wb.create_sheet("Sequences")
    ws_seq.append(['Visit #', 'Sequence #', 'Task #', 'Acquisition Mode', 'Data Volume (GB)', 'Over 1.8GB'])
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_seq[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Create set of flagged sequences
    flagged = {(alert.visit_id, alert.obs_id, alert.mode) for alert in alerts}
    
    # Add sequence data
    for visit in visits:
        for seq in visit.observation_sequences:
            for mode, vol_bytes in seq.data_volumes.items():
                vol_gb = vol_bytes / (1e9)
                is_flagged = (visit.visit_id, seq.obs_id, mode) in flagged
                ws_seq.append([
                    visit.visit_id,
                    seq.obs_id,
                    seq.task_id,
                    mode,
                    vol_gb,
                    "Yes" if is_flagged else "No"
                ])
    
    # Auto-adjust column widths
    ws_seq.column_dimensions['A'].width = 12
    ws_seq.column_dimensions['B'].width = 12
    ws_seq.column_dimensions['C'].width = 12
    ws_seq.column_dimensions['D'].width = 35
    ws_seq.column_dimensions['E'].width = 18
    ws_seq.column_dimensions['F'].width = 12
    
    # Format data volume column
    for row in ws_seq.iter_rows(min_row=2, max_row=ws_seq.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '0.000000'
    
    # ===== SHEET 2: VISITS =====
    ws_visits = wb.create_sheet("Visits")
    ws_visits.append(['Visit #', 'Total Data Volume (GB)'])
    
    # Style header
    for cell in ws_visits[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add visit data
    for visit in visits:
        total_bytes = sum(visit.total_data_volumes.values())
        total_gb = total_bytes / (1e9)
        ws_visits.append([visit.visit_id, total_gb])
    
    # Auto-adjust column widths
    ws_visits.column_dimensions['A'].width = 12
    ws_visits.column_dimensions['B'].width = 25
    
    # Format data volume column
    for row in ws_visits.iter_rows(min_row=2, max_row=ws_visits.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0.000000'
    
    # ===== SHEET 3: TASKS =====
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append(['Task #', 'Total Data Volume (GB)'])
    
    # Style header
    for cell in ws_tasks[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Aggregate by task_id (first 4 digits)
    task_volumes = {}
    for visit in visits:
        for seq in visit.observation_sequences:
            task_id = seq.task_id
            if task_id not in task_volumes:
                task_volumes[task_id] = 0.0
            for vol_bytes in seq.data_volumes.values():
                task_volumes[task_id] += vol_bytes / (1e9)
    
    # Add sorted task data (numerically if possible)
    for task_id in sorted(task_volumes.keys(), key=lambda x: int(x) if x and x.isdigit() else 0):
        ws_tasks.append([task_id, task_volumes[task_id]])
    
    # Auto-adjust column widths
    ws_tasks.column_dimensions['A'].width = 12
    ws_tasks.column_dimensions['B'].width = 25
    
    # Format data volume column
    for row in ws_tasks.iter_rows(min_row=2, max_row=ws_tasks.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '0.000000'
    
    # Save workbook
    wb.save(output_file)
    print(f"✓ Exported all data to {output_file}")
    print(f"  - Sheet 1: Sequences ({ws_seq.max_row - 1} rows)")
    print(f"  - Sheet 2: Visits ({ws_visits.max_row - 1} rows)")
    print(f"  - Sheet 3: Tasks ({ws_tasks.max_row - 1} rows)")

def print_data_volume_alerts(alerts: List[DataVolumeAlert]):
    """Print flagged high-volume sequences."""
    if not alerts:
        print("\n✓ No sequences exceed 1.8 GB threshold\n")
        return
    
    print(f"\n{'='*70}")
    print(f"⚠️  DATA VOLUME ALERTS ({len(alerts)} sequences flagged)")
    print(f"{'='*70}")
    for alert in sorted(alerts, key=lambda a: a.volume_gb, reverse=True):
        print(alert)
    print(f"{'='*70}\n")

def print_visit_summary(visit: VisitData):
    """Pretty print visit data and calculated volumes."""
    print(f"\n{'='*70}")
    print(f"VISIT {visit.visit_id} SUMMARY")
    print(f"{'='*70}")
    print(f"Total Observation Sequences: {len(visit.observation_sequences)}\n")
    
    for i, seq in enumerate(visit.observation_sequences, 1):
        print(f"Sequence {i}: {seq.obs_id}")
        print(f"  Target: {seq.target or 'N/A'}")
        print(f"  Acquisition Modes: {seq.modes}")
        
        if seq.vis_params:
            print(f"  VIS Parameters:")
            for key, val in sorted(seq.vis_params.items())[:]:  # Show first 5
                print(f"    {key}: {val}")
        
        if seq.nir_params:
            print(f"  NIR Parameters:")
            for key, val in sorted(seq.nir_params.items())[:]:  # Show first 5
                print(f"    {key}: {val}")
        
        print(f"  Data Volumes:")
        for mode, vol_bytes in seq.data_volumes.items():
            vol_mb = vol_bytes / (1e6)
            vol_gb = vol_bytes / (1e9)
            print(f"    {mode}: {vol_mb:.2f} MB ({vol_gb:.4f} GB)")
        print()
    
    print(f"\n{'='*70}")
    print(f"TOTAL DATA VOLUMES FOR VISIT {visit.visit_id}")
    print(f"{'='*70}")
    for mode, vol_bytes in visit.total_data_volumes.items():
        vol_mb = vol_bytes / (1e6)
        vol_gb = vol_bytes / (1e9)
        print(f"{mode:30s}: {vol_mb:12.2f} MB ({vol_gb:8.4f} GB)")
    
    total_all = sum(visit.total_data_volumes.values())
    total_gb = total_all / (1e9)
    print(f"{'-'*70}")
    print(f"{'TOTAL ALL MODES':30s}: {total_gb:8.4f} GB")
    print(f"{'='*70}\n")